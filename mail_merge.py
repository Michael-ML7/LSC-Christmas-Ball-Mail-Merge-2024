import smtplib, ssl, email
from email import encoders
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from mail_merge_body import email_body
from mail_merge_body_internal import email_body_internal

subject = "[Ticket Confirmation and Payment] Carpe Noctem - La Salle College Christmas Ball 2024"
sender = "lasallechristmasball2024@gmail.com"
password = "lpvd nlgk znkg iknx"

def send_email(subject, sender_email, recipient_email, password, recipient_name, recipient_ticket_type, recipient_price, souven):
    body = email_body(recipient_name, recipient_ticket_type, recipient_price, souven)
    mail_detail = MIMEMultipart("alternative")
    mail_detail['Subject'] = subject
    mail_detail['From'] = sender_email
    mail_detail['To'] = recipient_email

    msg = MIMEText(body, "html")
    mail_detail.attach(msg)
   
    with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp_server:
        smtp_server.login(sender_email, password)
        smtp_server.sendmail(sender_email, recipient_email, mail_detail.as_string())
    print("Message sent to " + recipient_name + " at " + recipient_email)

def send_email_internal(subject, sender_email, recipient_email, password, recipient_name, recipient_ticket_type, recipient_price):
    body = email_body_internal(recipient_name, recipient_ticket_type, recipient_price)
    mail_detail = MIMEMultipart("alternative")
    mail_detail['Subject'] = subject
    mail_detail['From'] = sender_email
    mail_detail['To'] = recipient_email

    msg = MIMEText(body, "html")
    mail_detail.attach(msg)
   
    with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp_server:
        smtp_server.login(sender_email, password)
        smtp_server.sendmail(sender_email, recipient_email, mail_detail.as_string())
    print("Message sent to " + recipient_name + " at " + recipient_email)

import openpyxl
wb = openpyxl.load_workbook('La Salle College Annual Christmas Ball 2024 - Carpe Noctem (Responses) (mail merge).xlsx', data_only=True)
ws = wb['Form Responses 1']
# wb = openpyxl.load_workbook('testing.xlsx')
# ws = wb['mail merge 1213']

for i in range(1, 353):
    if ws[f'X{i}'].value == "sent" or ws[f'X{i}'].value == "FAULT" or ws[f'X{i}'].value == "RESOLVED":
        continue
    ticket_type = ""
    price = ""
    if ws[f'E{i}'].value == "External (F3-6 female students)":
        if ws[f'R{i}'].value == "Yes":
            ticket_type = "Lady ticket (1 only)"
            price = "$200"
        if ws[f'S{i}'].value == "Yes":
            ticket_type = "Package for 4 ladies"
            price = "$640"
        if ws[f'T{i}'].value == "Yes":
            ticket_type = "Package for 6 ladies"
            price = "$900"
        souvenir = False
        if ws[f'W{i}'].value != None and ws[f'W{i}'].value != 0:
            souvenir = True
        print("External:", ws[f'N{i}'].value, ws[f'K{i}'].value, ticket_type, price, souvenir)
        # send_email(subject, sender, ws[f'N{i}'].value, password, ws[f'K{i}'].value, ticket_type, price, souvenir)
    else:
        if ws[f'P{i}'].value == "Yes":
            ticket_type = "Gentleman ticket (1 only)"
            price = "$200"
        if ws[f'Q{i}'].value == "Yes":
            ticket_type = "Couple ticket"
            price = "$340"
        student_email = ws[f'I{i}'].value + "@lsc.hk"
        print("Internal:", student_email, ws[f'F{i}'].value, ticket_type, price)
        # send_email_internal(subject, sender, student_email, password, ws[f'F{i}'].value, ticket_type, price)

# case handling
# send_email(subject, sender, "s20406@lsc.hk", password, "Michael Lam", "Package for 6 ladies", "$900", False)
# send_email(subject, sender, "s19294@lsc.hk", password, "Daniel Wong 6B", "Package for 6 ladies", "$900", False)
# send_email(subject, sender, "s19294@lsc.hk", password, "Daniel Wong 6B", "Package for 6 ladies", "$900", False)
# send_email(subject, sender, "tszyincheung88@gmail.com", password, "Cheung Tsz Yin", "Lady ticket (1 only)", "$200", False)
# send_email(subject, sender, "cchiu.0501@gmail.com", password, "Chiu Chrissy", "Package for 4 ladies", "$640", False)
# send_email_internal(subject, sender, "sisyphusprometheus@gmail.com", password, "Cheung Chun Lok", "Gentleman ticket (1 only)", "$200")
# send_email_internal(subject, sender, "s20175@lsc.hk", password, "OWW MY GAWDDDDD BREDNAN 😍😘❤", "Couple tickets ❤❤", "$340 ❤❤❤❤❤❤❤❤❤")
# send_email_internal(subject, sender, "s20406@lsc.hk", password, "Michael Lam", "Gentleman ticket (1 only)", "$200")